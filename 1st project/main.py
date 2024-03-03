from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox

#fn = 'data.xlsx'
#wb = load_workbook(fn)
#ws = wb['data']
# ws['A5'] = 'Привет, мир!'
# ws.append(['Один', 'Два', 'Три'])

'''
for row in range(1, 4):
    for col in range(1, 4):
        value = str(row) + str(col)
        cell = ws.cell(row=row, column=col)
        cell.value = value
'''

def save():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    data = (e.get(), lb['text'])
    ws.append(data)
    wb.save(fn)
    wb.close()
    messagebox.askokcancel('Сохранение', 'Успешно сохранено')


root = Tk()
root.title('Тест')
root.geometry('200x200')
root.resizable(0, 0)


e = Entry(root)
e.pack()
lb = Label(root, text='1 лейбл', font='Arial')
lb.pack()
btn = Button(root, text='Сохранить', font='Arial', command=save)
btn.pack()

